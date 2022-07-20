from xml.dom.minidom import parse
from tkinter import filedialog
from openpyxl.styles import PatternFill,Border,Side,colors
import openpyxl,os,time,requests,tkinter,you_get
from easygui import choicebox,enterbox
from tools import crc2uid
from win10toast import ToastNotifier
chats = []
toast = ToastNotifier()
if not os.path.exists('download'):
    os.mkdir('download')
def bulletchat(path,name):
    #打开文件
    xmlf = parse(path)
    data = xmlf.documentElement
    infos = data.getElementsByTagName('d')
    #读取信息
    for i in infos:
        chat = {}
        p = i.getAttribute('p').split(',')
        chat['time'] = float(p[0])
        chat['type'] = int(p[1])
        chat['size'] = int(p[2])
        chat['color'] = int(p[3])
        chat['timestamp'] = int(p[4])
        chat['pool'] = int(p[5])
        chat['uid_crc32'] = p[6]
        chat['text'] = i.childNodes[0].nodeValue
        chats.append(chat)
    chats.sort(key=lambda chat:chat['time'])
    #创建excel文件
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = name
    #设置表头
    sheet['A1'] = '时间点(单位：秒)'
    sheet['B1'] = '类型'
    sheet['C1'] = '字体大小(单位：像素)'
    sheet['D1'] = '颜色'
    sheet['E1'] = '发送时间'
    sheet['F1'] = '发送者(crc32加密)'
    sheet['G1'] = '弹幕内容'
    #加入内容
    for i in range(len(chats)):
        #时间点
        sheet['A'+str(i+2)] = chats[i]['time']
        #弹幕类型
        if chats[i]['pool'] == 0:
            if chats[i]['type'] == 1:
                sheet['B'+str(i+2)] = '滚动弹幕'
            elif chats[i]['type'] == 4:
                sheet['B'+str(i+2)] = '底部弹幕'
            elif chats[i]['type'] == 5:
                sheet['B'+str(i+2)] = '顶部弹幕'
            elif chats[i]['type'] == 6:
                sheet['B'+str(i+2)] = '逆向弹幕'
            elif chats[i]['type'] == 7:
                sheet['B'+str(i+2)] = '特殊弹幕'
        elif chats[i]['pool'] == 1:
            if chats[i]['type'] == 7:
                sheet['B'+str(i+2)] = '精确弹幕'
        elif chats[i]['pool'] == 2:
            if chats[i]['type'] == 9:
                sheet['B'+str(i+2)] = 'BAS弹幕'
        #字体大小
        sheet['C'+str(i+2)] = chats[i]['size']
        #弹幕颜色
        color_16 = hex(chats[i]['color']).replace('0x','')
        if len(color_16) < 6:
            color_16_l = list(color_16)
            for j in range(6-len(color_16)):
                color_16_l.insert(0,'0')
            color_16 = ''.join(color_16_l)
        fill_color = PatternFill(fill_type='solid',fgColor=color_16)
        sheet['D'+str(i+2)].fill = fill_color
        #发送时间
        sendtime = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(chats[i]['timestamp']))
        sheet['E'+str(i+2)] = sendtime
        #发送用户
        sheet['F'+str(i+2)] = chats[i]['uid_crc32']
        #内容
        sheet['G'+str(i+2)] = chats[i]['text']
    #单元格居中及添加框线
    alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center", text_rotation=0)
    border_set = Border(left=Side(style='thin', color=colors.BLACK),right=Side(style='thin', color=colors.BLACK),top=Side(style='thin', color=colors.BLACK),bottom=Side(style='thin', color=colors.BLACK))
    for i in sheet['A']:
        i.alignment = alignment
        i.border = border_set
    for i in sheet['B']:
        i.alignment = alignment
        i.border = border_set
    for i in sheet['C']:
        i.alignment = alignment
        i.border = border_set
    for i in sheet['D']:
        i.alignment = alignment
        i.border = border_set
    for i in sheet['E']:
        i.alignment = alignment
        i.border = border_set
    for i in sheet['F']:
        i.alignment = alignment
        i.border = border_set
    for i in sheet['G']:
        i.alignment = alignment
        i.border = border_set
    #调整列宽
    sheet.column_dimensions['A'].width = 17.375
    sheet.column_dimensions['B'].width = 13.0
    sheet.column_dimensions['C'].width = 21.5
    sheet.column_dimensions['D'].width = 5.25
    sheet.column_dimensions['E'].width = 21.625
    sheet.column_dimensions['F'].width = 18.75
    sheet.column_dimensions['G'].width = 44.25
    #存储
    book.save('download/'+name+' 弹幕信息.xlsx')
def getBulletChat(bv,p):
    toast.show_toast("BiliBili Tools","获取弹幕中...",duration=1)
    print('获取弹幕中...')
    #获取cid
    cid_result = requests.get('https://api.bilibili.com/x/player/pagelist?bvid='+bv).json()
    cid = cid_result['data'][p-1]['cid']
    #写入弹幕文件
    xml_b = requests.get('https://comment.bilibili.com/'+str(cid)+'.xml').content
    xml_f = open('download/tmp.xml','wb')
    xml_f.write(xml_b)
    xml_f.close()
def getName(bv,p):
    toast.show_toast("BiliBili Tools","获取视频标题中...",duration=1)
    print('获取视频标题中...')
    #获取视频标题
    title_result = requests.get('https://api.szfx.top/bilibili/api.php?bv='+bv).json()
    name = title_result['title']
    #如果是分p的还要进一步改名
    #获取分p名
    pname_result = requests.get('https://api.bilibili.com/x/player/pagelist?bvid='+bv).json()
    if len(pname_result['data']) != 1:
        name = name + ' (P'+str(p)+'. '+pname_result['data'][p-1]['part']+')'
    name = name.replace('/','-').replace('[','(').replace(']',')')
    return name
def downloadVideo(bv,p,name):
    toast.show_toast("BiliBili Tools","下载视频中...",duration=1)
    print('下载视频中...')
    if not os.path.exists('download/'+name+'.mp4'):
        os.system('you-get -o download https://www.bilibili.com/video/'+bv+'?p='+str(p))
        os.system(r'del "download\\'+name+'.cmt.xml"')
        #如果是flv格式的直接改成mp4格式的
        if os.path.exists('download/'+name+'.flv'):
            os.chdir('download')
            os.system('ren "'+name+'.flv" "'+name+'.mp4"')
            os.chdir('..')
    else:
        toast.show_toast("BiliBili Tools","文件已存在",duration=1)
        print('文件已存在')
#----------------------------------------------------------------
#隐藏窗口
win = tkinter.Tk()
win.withdraw()
#选择模式
mode = choicebox("请选择工作模式","工作模式",['下载视频和弹幕(单独存放)，并进行弹幕分析','下载视频并嵌入弹幕，并进行弹幕分析','将本地弹幕嵌入本地视频，并分析弹幕','仅分析弹幕','分析本地弹幕','解密用户uid(被CRC32加密)'])
if mode == '下载视频和弹幕(单独存放)，并进行弹幕分析':
    #输入BV号
    bv = enterbox("请输入视频的BV号：","BV号")
    #输入分集
    part = enterbox("请输入视频分集的集数(留空代表第一集)：","集数")
    if part == None or part == '':
        p = 1
    else:
        p = int(part)
    #获取视频名
    name = getName(bv,p)
    #下载视频及弹幕
    downloadVideo(bv,p,name)
    getBulletChat(bv,p)
    #分析弹幕
    toast.show_toast("BiliBili Tools","分析弹幕中...",duration=1)
    print('分析弹幕中...')
    if not os.path.exists('download/'+name+' 弹幕信息.xlsx'):
        bulletchat('download/tmp.xml',name)
    else:
        toast.show_toast("BiliBili Tools","文件已存在",duration=1)
        print('文件已存在')
    #播放
    os.system(r'del download\\tmp.xml')
    print('程序执行完毕！')
elif mode == '下载视频并嵌入弹幕，并进行弹幕分析':
    #输入BV号
    bv = enterbox("请输入视频的BV号：","BV号")
    #输入分集
    part = enterbox("请输入视频分集的集数(留空代表第一集)：","集数")
    if part == None or part == '':
        p = 1
    else:
        p = int(part)
    #获取视频标题
    name = getName(bv,p)
    #下载视频及弹幕
    downloadVideo(bv,p,name)
    getBulletChat(bv,p)
    #分析弹幕
    toast.show_toast("BiliBili Tools","分析弹幕中...",duration=1)
    print('分析弹幕中...')
    if not os.path.exists('download/'+name+' 弹幕信息.xlsx'):
        bulletchat('download/tmp.xml',name)
    else:
        toast.show_toast("BiliBili Tools","文件已存在",duration=1)
        print('文件已存在')
    #转换弹幕
    os.chdir('download')
    toast.show_toast("BiliBili Tools","转换弹幕中...",duration=1)
    print("转换弹幕中...")
    if not os.path.exists('tmp.ass'):
        os.system(r'..\ffmpeg\xml2ass.exe tmp.xml')
    else:
        toast.show_toast("BiliBili Tools","文件已存在",duration=1)
        print('文件已存在')
    #嵌入弹幕
    toast.show_toast("BiliBili Tools","嵌入弹幕中...",duration=1)
    print("嵌入弹幕中...")
    if not os.path.exists(name+'(弹幕嵌入).mp4'):
        os.system(r'..\ffmpeg\ffmpeg.exe -i "'+name+'.mp4" -vf subtitles=tmp.ass -vcodec libx264 "'+name+'(弹幕嵌入).mp4"')
    else:
        toast.show_toast("BiliBili Tools","文件已存在",duration=1)
        print('文件已存在')
    #删除临时文件
    toast.show_toast("BiliBili Tools","删除临时文件中...",duration=1)
    print("删除临时文件中...")
    os.system('del tmp.ass')
    #播放
    os.chdir('..')
    os.system(r'del download\\tmp.xml')
    print("程序执行完毕！")
elif mode == '将本地弹幕嵌入本地视频，并分析弹幕':
    chatpath = filedialog.askopenfilename(title='请选择弹幕文件',filetype=[('XML File', '.xml')])
    videopath = filedialog.askopenfilename(title='请选择视频文件',filetype=[('MP4 File','.mp4')])
    name = os.path.splitext(os.path.split(videopath)[1])[0]
    chatname = os.path.splitext(os.path.split(chatpath)[1])[0]
    #分析弹幕
    toast.show_toast("BiliBili Tools","分析弹幕中...",duration=1)
    print('分析弹幕中...')
    if not os.path.exists('download/'+name+' 弹幕信息.xlsx'):
        bulletchat(chatpath,name)
    else:
        toast.show_toast("BiliBili Tools","文件已存在",duration=1)
        print('文件已存在')
    #复制弹幕及视频
    toast.show_toast("BiliBili Tools","复制弹幕及视频中...",duration=1)
    print('复制弹幕及视频中...')
    os.chdir('download')
    cmdchat = 'copy "'+chatpath+'" tmp.xml'
    cmdchat = cmdchat.replace('/',r'\\')
    cmdvideo = 'copy "'+videopath+'" tmp.mp4'
    cmdvideo = cmdvideo.replace('/',r'\\')
    os.system(cmdchat)
    os.system(cmdvideo)
    #转换弹幕
    toast.show_toast("BiliBili Tools","转换弹幕中...",duration=1)
    print("转换弹幕中...")
    if not os.path.exists('tmp.ass'):
        os.system(r'..\ffmpeg\xml2ass.exe tmp.xml')
    else:
        toast.show_toast("BiliBili Tools","文件已存在",duration=1)
        print('文件已存在')
    #嵌入弹幕
    toast.show_toast("BiliBili Tools","嵌入弹幕中...",duration=1)
    print("嵌入弹幕中...")
    if not os.path.exists(name+'(弹幕嵌入).mp4'):
        os.system(r'..\ffmpeg\ffmpeg.exe -i tmp.mp4 -vf subtitles=tmp.ass -vcodec libx264 "'+name+'(弹幕嵌入).mp4"')
    else:
        toast.show_toast("BiliBili Tools","文件已存在",duration=1)
        print('文件已存在')
    #删除临时文件
    toast.show_toast("BiliBili Tools","删除临时文件中...",duration=1)
    print("删除临时文件中...")
    os.system('del tmp.ass')
    os.system('del tmp.mp4')
    #播放
    os.chdir('..')
    os.system(r'del download\\tmp.xml')
    print('程序执行完毕！')
elif mode == '仅分析弹幕':
    #输入BV号
    bv = enterbox("请输入视频的BV号：","BV号")
    #输入分集
    part = enterbox("请输入视频分集的集数(留空代表第一集)：","集数")
    if part == None or part == '':
        p = 1
    else:
        p = int(part)
    #获取标题
    name = getName(bv,p)
    #获取弹幕
    getBulletChat(bv,p)
    #分析弹幕
    toast.show_toast("BiliBili Tools","分析弹幕中...",duration=1)
    print('分析弹幕中...')
    if not os.path.exists('download/'+name+' 弹幕信息.xlsx'):
        bulletchat('download/tmp.xml',name)
    else:
        toast.show_toast("BiliBili Tools","文件已存在",duration=1)
        print('文件已存在')
    os.system(r'del download\\tmp.xml')
elif mode == '分析本地弹幕':
    #选择文件
    path = filedialog.askopenfilename(title='请选择弹幕文件',filetype=[('XML File', '.xml')])
    pather = os.path.split(path)[0] + '/'
    name = os.path.splitext(os.path.split(path)[1])[0]
    #分析弹幕
    toast.show_toast("BiliBili Tools","分析弹幕中...",duration=1)
    print('分析弹幕中...')
    if not os.path.exists('download/'+name+' 弹幕信息.xlsx'):
        bulletchat(path,name)
    else:
        toast.show_toast("BiliBili Tools","文件已存在",duration=1)
        print('文件已存在')
elif mode == '解密用户uid(被CRC32加密)':
    crc_uid = enterbox("请输入被crc32加密的用户uid：","输入加密数据")
    uid = crc2uid.crc2uid(crc_uid)
    print(uid)
    #获取uid信息
    user_result = requests.get('https://tenapi.cn/bilibili/?uid='+uid).json()
    user_name = user_result['data']['name']
    user_level = user_result['data']['level']
    user_sex = user_result['data']['sex']
    user_description = user_result['data']['description']
    text_toast = 'uid：{}\n名称：{}\n等级：{}\n性别：{}'.format(uid,user_name,user_level,user_sex)
    text = "-----------用户信息-----------\nuid：{0}\n主页：https://space.bilibili.com/{0}\n名称：{1}\n等级：Lv.{2}\n性别：{3}\n简介：{4}\n------------------------------".format(uid,user_name,user_level,user_sex,user_description)
    '''
    print("-----------用户信息-----------")
    print("uid：",uid)
    print("主页：https://space.bilibili.com/"+uid)
    print("名称："+user_name)
    print("等级：Lv."+str(user_level))
    print("性别："+user_sex)
    print("简介："+user_description)
    print("------------------------------")
    '''
    print(text)
    toast.show_toast("用户信息",text_toast,duration=10)
