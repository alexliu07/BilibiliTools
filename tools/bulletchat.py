import time
from xml.dom.minidom import parse

import openpyxl
from openpyxl.styles import PatternFill, Border, Side, colors


def bulletchat(path, name):
    # 打开文件
    xmlf = parse(path)
    data = xmlf.documentElement
    infos = data.getElementsByTagName('d')
    # 读取信息
    for i in infos:
        chat = {}
        p = i.getAttribute('p').split(',')
        chat['time'] = float(p[0])
        chat['type'] = int(p[1])
        chat['size'] = int(p[2])
        chat['color'] = int(p[3])
        chat['timestamp'] = int(p[4])
        chat['pool'] = int(p[5])
        chat['uid_crc32'] = p[6]
        chat['text'] = i.childNodes[0].nodeValue
        chats.append(chat)
    chats.sort(key=lambda chat: chat['time'])
    # 创建excel文件
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = name
    # 设置表头
    sheet['A1'] = '时间点(单位：秒)'
    sheet['B1'] = '类型'
    sheet['C1'] = '字体大小(单位：像素)'
    sheet['D1'] = '颜色'
    sheet['E1'] = '发送时间'
    sheet['F1'] = '发送者(crc32加密)'
    sheet['G1'] = '弹幕内容'
    # 加入内容
    for i in range(len(chats)):
        # 时间点
        sheet['A' + str(i + 2)] = chats[i]['time']
        # 弹幕类型
        if chats[i]['pool'] == 0:
            if chats[i]['type'] == 1:
                sheet['B' + str(i + 2)] = '滚动弹幕'
            elif chats[i]['type'] == 4:
                sheet['B' + str(i + 2)] = '底部弹幕'
            elif chats[i]['type'] == 5:
                sheet['B' + str(i + 2)] = '顶部弹幕'
            elif chats[i]['type'] == 6:
                sheet['B' + str(i + 2)] = '逆向弹幕'
            elif chats[i]['type'] == 7:
                sheet['B' + str(i + 2)] = '特殊弹幕'
        elif chats[i]['pool'] == 1:
            if chats[i]['type'] == 7:
                sheet['B' + str(i + 2)] = '精确弹幕'
        elif chats[i]['pool'] == 2:
            if chats[i]['type'] == 9:
                sheet['B' + str(i + 2)] = 'BAS弹幕'
        # 字体大小
        sheet['C' + str(i + 2)] = chats[i]['size']
        # 弹幕颜色
        color_16 = hex(chats[i]['color']).replace('0x', '')
        if len(color_16) < 6:
            color_16_l = list(color_16)
            for j in range(6 - len(color_16)):
                color_16_l.insert(0, '0')
            color_16 = ''.join(color_16_l)
        fill_color = PatternFill(fill_type='solid', fgColor=color_16)
        sheet['D' + str(i + 2)].fill = fill_color
        # 发送时间
        sendtime = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(chats[i]['timestamp']))
        sheet['E' + str(i + 2)] = sendtime
        # 发送用户
        sheet['F' + str(i + 2)] = chats[i]['uid_crc32']
        # 内容
        sheet['G' + str(i + 2)] = chats[i]['text']
    # 单元格居中及添加框线
    alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center", text_rotation=0)
    border_set = Border(left=Side(style='thin', color=colors.BLACK), right=Side(style='thin', color=colors.BLACK),
                        top=Side(style='thin', color=colors.BLACK), bottom=Side(style='thin', color=colors.BLACK))
    for i in sheet['A']:
        i.alignment = alignment
        i.border = border_set
    for i in sheet['B']:
        i.alignment = alignment
        i.border = border_set
    for i in sheet['C']:
        i.alignment = alignment
        i.border = border_set
    for i in sheet['D']:
        i.alignment = alignment
        i.border = border_set
    for i in sheet['E']:
        i.alignment = alignment
        i.border = border_set
    for i in sheet['F']:
        i.alignment = alignment
        i.border = border_set
    for i in sheet['G']:
        i.alignment = alignment
        i.border = border_set
    # 调整列宽
    sheet.column_dimensions['A'].width = 17.375
    sheet.column_dimensions['B'].width = 13.0
    sheet.column_dimensions['C'].width = 21.5
    sheet.column_dimensions['D'].width = 5.25
    sheet.column_dimensions['E'].width = 21.625
    sheet.column_dimensions['F'].width = 18.75
    sheet.column_dimensions['G'].width = 44.25
    # 存储
    book.save('download/' + name + ' 弹幕信息.xlsx')
